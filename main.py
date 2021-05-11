import requests
import argparse
import json
import csv
import unicodedata

import os
import sys

import openpyxl

if sys.version_info >= (3, 0):
    import configparser as ConfigParser
else:
    import ConfigParser

from glob import glob
import ntpath

import time


class spotifyApp():
    def __init__(self):
        self.accessToken = ""
        self.client_id = '8a14ebc91cf34265b1acbfb777b946ab'
        self.client_secret = 'fe7ad45553294aff829b4a101f059fb5'
        self.scope = 'playlist-modify-public playlist-modify-private playlist-read-collaborative'
        self.refresh_token = 'AQAsQJRqXNVHkMsfAtvKciUg8LuinmLEi92NnDXDk2ePwsIJMc14Qci6Lhk7F-oAXOkOIcTnSJOOdMumR4yOMLVX8IbmRCzxIKZJ7jgytLwxAaEfYcY6pZMURNfCINJfwxs'
        self.excelheader = [
            'Year',
            'Artist',
            'Album',
            'Url',
            'Popularity'
        ]

    def get_access_token(self):
        url = 'https://accounts.spotify.com/api/token'
        payload = {
            'grant_type': 'refresh_token',
            'refresh_token': self.refresh_token
        }
        auth = (self.client_id, self.client_secret)
        token = requests.post(url, data=payload, auth=auth).json()
        self.accessToken = token['access_token']
        return

    def getData(self, inputfilenames, outputFolder):
        # csvfile = open(outputpath, 'w', newline='\n')#, encoding="utf-8-sig")#utf-8-sig, utf8
        # writer = csv.DictWriter(csvfile, delimiter=",", fieldnames=self.excelheader)
        # writer.writeheader()

        for inputFile in inputfilenames:
            self.get_access_token()
            outputfile = outputFolder + ntpath.basename(inputFile)
            print("outputfile:{}".format(outputfile))
            trackiddata = []
            year = 1900
            wb_inp = openpyxl.load_workbook(inputFile)
            sheet_inp = wb_inp.active
            row_inp = sheet_inp.max_row + 1
            for row in range(row_inp):
                if row < 2:
                    continue

                trackiddata.append(str(sheet_inp.cell(row=row, column=3).value).replace(
                    'spotify:track:', '').replace('https://open.spotify.com/track/', '').strip())
                year = sheet_inp.cell(row=row, column=14).value

            wb = openpyxl.Workbook()
            sheet = wb.active
            for i in range(5):
                sheet.cell(row=1, column=i + 1).value = self.excelheader[i]
            grow = 2

            for trackid in trackiddata:
                print('------------{}------------'.format(grow - 1))
                print("----------trackid:{}".format(trackid))

                while True:
                    try:
                        # get Popularity
                        pheaders = {'Origin': 'https://open.spotify.com',
                                    'Accept-Encoding': 'gzip, deflate, br',
                                    'Accept-Language': 'en',
                                    'Authorization': 'Bearer ' + self.accessToken,
                                    'Accept': 'application/json',
                                    # 'Referer': 'https://open.spotify.com/search/albums/year^%^3A1980',
                                    'Authority': 'api.spotify.com',
                                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.100 Safari/537.36'}
                        varhref = 'https://api.spotify.com/v1/tracks/{}'.format(trackid)

                        presponse = requests.get(
                            varhref, headers=pheaders).json()

                        albumid = presponse['album']['id']
                        print('----------albumid:{}'.format(albumid))

                        headers = {'Origin': 'https://open.spotify.com',
                                'Accept-Encoding': 'gzip, deflate, br',
                                'Accept-Language': 'en',
                                'Authorization': 'Bearer ' + self.accessToken,
                                'Accept': 'application/json',
                                # 'Referer': 'https://open.spotify.com/search/albums/year^%^3A1980',
                                'Authority': 'api.spotify.com',
                                'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.100 Safari/537.36'}

                        response = requests.get('https://api.spotify.com/v1/albums/{}'.format(albumid), headers=pheaders).json()

                        try:
                            varpopularity = response['popularity']
                        except:
                            varpopularity = 0

                        try:
                            varpopularity = response['popularity']
                        except:
                            varpopularity = 0

                        try:
                            varartist = response['artists'][0]['name']
                        except:
                            varartist = ''

                        try:
                            varalbum = response['name']
                        except:
                            varalbum = ''

                        try:
                            varurl = response['external_urls']['spotify']
                        except:
                            varurl = ''

                        print("Year:{}, Artist:{}, Album:{}, Url:{}, Popularity:{}".format(
                            year, varartist, varalbum, varurl, varpopularity))

                        sheet.cell(row=grow, column=1).value = year
                        sheet.cell(row=grow, column=2).value = varartist
                        sheet.cell(row=grow, column=3).value = varalbum
                        sheet.cell(row=grow, column=4).value = varurl
                        sheet.cell(
                            row=grow, column=5).value = varpopularity
                        grow += 1

                    except:
                        print("Request Error!")
                        time.sleep(1)
                        continue
                    break


            wb.save(outputfile)


def load_config():
    defaults = {
        'input': '',
        'output': ''
    }
    _settings_dir = "./"
    config_file = os.path.join(_settings_dir, "config.ini")
    if os.path.exists(config_file):
        print('Existing config.ini')
        try:
            # config = ConfigParser.SafeConfigParser()
            config = ConfigParser.ConfigParser()
            config.read(config_file)
            if config.has_section("global"):
                config_items = dict(config.items("global"))

                defaults['input'] = config_items['input']
                defaults['output'] = config_items['output']
        except ConfigParser.Error as e:
            print("\nError parsing config file: " + config_file)
            print(str(e))
            exit(1)

    return defaults


def getInputFileList(inputFolder):
    inputFilepaths = glob(inputFolder+"*.xlsx")
    print(inputFilepaths)
    return inputFilepaths


def startProcess():
    config_option = load_config()

    inputFolder = config_option['input']
    outputFolder = config_option['output']

    print('inputFolder:{}'.format(inputFolder))
    print('outputFolder:{}'.format(outputFolder))

    inputfilenames = getInputFileList(inputFolder)
    app = spotifyApp()
    app.getData(inputfilenames, outputFolder)


if __name__ == "__main__":
    startProcess()
    print("------------finish------------")
